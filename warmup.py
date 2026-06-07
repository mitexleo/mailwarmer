#!/usr/bin/env python3
"""
Mail server warm-up script
Reads recipients from an xlsx file and sends emails gradually over days.

Usage:
  Day 1:  python3 warmup.py --day 1
  Day 2:  python3 warmup.py --day 2
  ...etc

Or run continuously (waits between days):
  python3 warmup.py --auto

Requirements:
  pip install openpyxl python-dotenv
  cp .env.example .env    # then edit .env with your real config
"""

import smtplib
import time
import argparse
import logging
import json
import math
import os
import re
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables from .env (ignored by git)
load_dotenv()

# ── Config (from .env) ────────────────────────────────────────────────────────

SMTP_HOST = os.getenv("SMTP_HOST")
SMTP_PORT = int(os.getenv("SMTP_PORT", "25"))
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASS = os.getenv("SMTP_PASS")
USE_TLS   = os.getenv("USE_TLS", "false").lower() == "true"

FROM_NAME  = os.getenv("FROM_NAME")
FROM_EMAIL = os.getenv("FROM_EMAIL")

XLSX_FILE      = os.getenv("XLSX_FILE")
STATE_FILE     = os.getenv("STATE_FILE", "warmup_state.json")
EMAIL_SUBJECT  = os.getenv("EMAIL_SUBJECT")
TEXT_TEMPLATE  = os.getenv("TEXT_TEMPLATE", "email_body.txt.template")

WARMUP_DAYS           = int(os.getenv("WARMUP_DAYS", "14"))
DELAY_BETWEEN_EMAILS  = int(os.getenv("DELAY_BETWEEN_EMAILS", "10"))
DELAY_BETWEEN_DAYS    = int(os.getenv("DELAY_BETWEEN_DAYS", "86400"))

required_vars = {
    "SMTP_HOST": SMTP_HOST,
    "SMTP_USER": SMTP_USER,
    "SMTP_PASS": SMTP_PASS,
    "FROM_NAME": FROM_NAME,
    "FROM_EMAIL": FROM_EMAIL,
    "XLSX_FILE": XLSX_FILE,
    "EMAIL_SUBJECT": EMAIL_SUBJECT,
}
missing = [name for name, val in required_vars.items() if not val]
if missing:
    raise SystemExit(
        f"Missing required env vars: {', '.join(missing)}\n"
        "Copy .env.example to .env and fill in your values."
    )

# ── Logging setup ────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("warmup.log"),
    ]
)
log = logging.getLogger(__name__)


# ── Load template files ──────────────────────────────────────────────────────

def _load_template(path):
    """Load a template file and replace {{FROM_NAME}} placeholders."""
    try:
        with open(path) as f:
            content = f.read()
        return content.replace("{{FROM_NAME}}", FROM_NAME)
    except FileNotFoundError:
        log.warning("Template file '%s' not found, using empty body.", path)
        return ""


def _text_to_html(text):
    """Convert plain text into a basic HTML email body."""
    paragraphs = text.strip().split("\n\n")
    html_paragraphs = []
    for p in paragraphs:
        lines = p.replace("\n", "<br>")
        html_paragraphs.append(f"<p>{lines}</p>")
    body = "\n".join(html_paragraphs)
    return f"""\
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
</head>
<body style="font-family: Arial, sans-serif; padding: 20px;">
{body}
</body>
</html>
"""


# ── Email content ─────────────────────────────────────────────────────────────

SUBJECT    = EMAIL_SUBJECT
TEXT_SOURCE = _load_template(TEXT_TEMPLATE)
HTML_BODY   = _text_to_html(TEXT_SOURCE) if TEXT_SOURCE else ""

# Persist the generated HTML for inspection
if HTML_BODY:
    with open("email_body.html", "w") as f:
        f.write(HTML_BODY)
    log.info("Generated email_body.html from %s", TEXT_TEMPLATE)


# ── Schedule generator ────────────────────────────────────────────────────────

def build_schedule(total_recipients, days):
    """
    Generate a warm-up schedule that ramps up gradually over `days`.
    Returns a dict: day_number → emails_to_send.
    """
    if days <= 0 or total_recipients <= 0:
        return {}

    base = total_recipients // days
    remainder = total_recipients % days

    schedule = {}
    for d in range(1, days + 1):
        # Linear ramp: start small, grow toward peak
        ratio = (d / days)
        raw = max(1, round(total_recipients * ratio * (2 / (days + 1)) * d))
        # Clamp: at least 1, at most remaining
        already = sum(schedule.get(i, 0) for i in range(1, d))
        remaining = total_recipients - already
        quota = min(raw, remaining)
        # If this is the last day, send whatever is left
        if d == days:
            quota = remaining
        schedule[d] = quota

    # Final sanity: ensure totals match
    total = sum(schedule.values())
    if total < total_recipients:
        schedule[days] += total_recipients - total
    elif total > total_recipients:
        schedule[days] -= total - total_recipients

    # Remove zero-quota days
    return {d: q for d, q in schedule.items() if q > 0}


# ── Helpers ───────────────────────────────────────────────────────────────────



def load_recipients():
    wb = load_workbook(XLSX_FILE, read_only=True)
    ws = wb.active
    return [row[0] for row in ws.iter_rows(values_only=True) if row[0]]


def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE) as f:
            return json.load(f)
    return {"sent_index": 0, "last_day": 0}


def save_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)


def build_message(to_email):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = SUBJECT
    msg["From"]    = f"{FROM_NAME} <{FROM_EMAIL}>"
    msg["To"]      = to_email
    # Always send HTML-only (no text/plain fallback)
    if HTML_BODY:
        msg.attach(MIMEText(HTML_BODY, "html"))
    return msg


def send_batch(recipients, count, state):
    start = state["sent_index"]
    batch = recipients[start:start + count]

    if not batch:
        log.info("No more recipients to send to.")
        return 0

    log.info(f"Connecting to {SMTP_HOST}:{SMTP_PORT} ...")
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
            smtp.ehlo()
            if USE_TLS:
                smtp.starttls()
                smtp.ehlo()
            smtp.login(SMTP_USER, SMTP_PASS)
            log.info("Authenticated OK.")

            sent = 0
            for email in batch:
                try:
                    msg = build_message(email)
                    smtp.sendmail(FROM_EMAIL, [email], msg.as_string())
                    log.info(f"  ✓ Sent to {email}")
                    sent += 1
                    state["sent_index"] += 1
                    save_state(state)
                    if email != batch[-1]:
                        time.sleep(DELAY_BETWEEN_EMAILS)
                except smtplib.SMTPException as e:
                    log.warning(f"  ✗ Failed {email}: {e}")

    except Exception as e:
        log.error(f"SMTP connection error: {e}")
        return 0

    return sent


# ── Main ──────────────────────────────────────────────────────────────────────

def run_day(day, recipients, state, schedule):
    quota = schedule.get(day)
    if quota is None:
        log.error(f"No schedule entry for day {day}. Check WARMUP_DAYS config.")
        return

    log.info(f"=== Day {day}: sending up to {quota} emails (sent so far: {state['sent_index']}) ===")
    sent = send_batch(recipients, quota, state)
    state["last_day"] = day
    save_state(state)
    log.info(f"=== Day {day} done: {sent} emails sent. Total sent: {state['sent_index']} ===")


def main():
    parser = argparse.ArgumentParser(description="SMTP warm-up script")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--day", type=int, help="Run a specific day")
    group.add_argument("--auto", action="store_true", help="Run all days automatically with configured delays")
    args = parser.parse_args()

    recipients = load_recipients()
    log.info(f"Loaded {len(recipients)} recipients from {XLSX_FILE}")

    schedule = build_schedule(len(recipients), WARMUP_DAYS)
    log.info(f"Auto-generated schedule over {len(schedule)} days: {dict(schedule)}")

    state = load_state()
    log.info(f"Resuming: sent_index={state['sent_index']}, last_day={state['last_day']}")

    if args.day:
        run_day(args.day, recipients, state, schedule)
    elif args.auto:
        start_day = state["last_day"] + 1
        for day in range(start_day, max(schedule.keys()) + 1):
            run_day(day, recipients, state, schedule)
            if day < max(schedule.keys()) and state["sent_index"] < len(recipients):
                log.info(f"Sleeping {DELAY_BETWEEN_DAYS}s before day {day + 1} ...")
                time.sleep(DELAY_BETWEEN_DAYS)


if __name__ == "__main__":
    main()
