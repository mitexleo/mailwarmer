"""
Core SMTP warm-up logic — load config, recipients, schedule,
send emails, and track state.  Used by both the CLI and the GUI.
"""

import csv
import json
import logging
import os
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from dotenv import load_dotenv

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# ── Version (used for self-update check) ──────────────────────────────────────

__version__ = "1.1.0"


# ── Config ────────────────────────────────────────────────────────────────────

def load_config(env_path):
    """Load .env file and return a config dict."""
    load_dotenv(dotenv_path=env_path, override=True)

    cfg = {
        "smtp_host":      os.getenv("SMTP_HOST"),
        "smtp_port":      int(os.getenv("SMTP_PORT", "25")),
        "smtp_user":      os.getenv("SMTP_USER"),
        "smtp_pass":      os.getenv("SMTP_PASS"),
        "use_tls":        os.getenv("USE_TLS", "false").lower() == "true",
        "from_name":      os.getenv("FROM_NAME"),
        "from_email":     os.getenv("FROM_EMAIL"),
        "email_subject":  os.getenv("EMAIL_SUBJECT"),
        "warmup_days":    int(os.getenv("WARMUP_DAYS", "14")),
        "delay_emails":   int(os.getenv("DELAY_BETWEEN_EMAILS", "10")),
        "delay_days":     int(os.getenv("DELAY_BETWEEN_DAYS", "86400")),
        "state_file":     os.getenv("STATE_FILE", "warmup_state.json"),
    }

    required = ["smtp_host", "smtp_user", "smtp_pass", "from_name", "from_email", "email_subject"]
    missing = [k.upper() for k in required if not cfg[k]]
    if missing:
        raise ValueError(f"Missing required env vars: {', '.join(missing)}")

    return cfg


# ── Recipients ────────────────────────────────────────────────────────────────

def load_recipients(path):
    """Load email addresses from an XLSX or CSV file (first column)."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Data file not found: {path}")

    ext = os.path.splitext(path)[1].lower()

    if ext == ".xlsx":
        if load_workbook is None:
            raise ImportError("openpyxl is required for .xlsx files — pip install openpyxl")
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        return [str(row[0]).strip() for row in ws.iter_rows(values_only=True) if row[0]]

    elif ext == ".csv":
        recipients = []
        with open(path, newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                if row and row[0].strip():
                    recipients.append(row[0].strip())
        return recipients

    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .xlsx or .csv.")


# ── HTML ──────────────────────────────────────────────────────────────────────

def load_html_body(path):
    """Read and return the HTML email body from a file."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Email HTML file not found: {path}")
    with open(path) as f:
        return f.read()


# ── Schedule ──────────────────────────────────────────────────────────────────

def build_schedule(total, days):
    """Generate a gradual ramp-up schedule over N days."""
    if days <= 0 or total <= 0:
        return {}

    schedule = {}
    for d in range(1, days + 1):
        if d == days:
            already = sum(schedule.values())
            schedule[d] = total - already
        else:
            ratio = d / days
            raw = max(1, round(total * ratio * (2 / (days + 1)) * d))
            already = sum(schedule.values())
            remaining = total - already
            schedule[d] = min(raw, remaining)

    return {d: q for d, q in schedule.items() if q > 0}


# ── State ─────────────────────────────────────────────────────────────────────

def load_state(path):
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {"sent_index": 0, "last_day": 0}


def save_state(path, state):
    with open(path, "w") as f:
        json.dump(state, f, indent=2)


# ── Sender ────────────────────────────────────────────────────────────────────

def build_message(to_email, cfg, html_body):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = cfg["email_subject"]
    msg["From"] = f'{cfg["from_name"]} <{cfg["from_email"]}>'
    msg["To"] = to_email
    msg.attach(MIMEText(html_body, "html"))
    return msg


def send_batch(recipients, count, state, cfg, html_body, log, progress_callback=None):
    """Send up to *count* emails starting from state['sent_index'].

    Calls ``progress_callback(current, total)`` after each email if provided.
    Returns the number of successfully sent emails.
    """
    start = state["sent_index"]
    batch = recipients[start:start + count]
    total_in_batch = len(batch)

    if not batch:
        log.info("No more recipients to send to.")
        return 0

    log.info("Connecting to %s:%s ...", cfg["smtp_host"], cfg["smtp_port"])
    try:
        with smtplib.SMTP(cfg["smtp_host"], cfg["smtp_port"], timeout=30) as smtp:
            smtp.ehlo()
            if cfg["use_tls"]:
                smtp.starttls()
                smtp.ehlo()
            smtp.login(cfg["smtp_user"], cfg["smtp_pass"])
            log.info("Authenticated OK.")

            sent = 0
            for idx, email in enumerate(batch):
                try:
                    msg = build_message(email, cfg, html_body)
                    smtp.sendmail(cfg["from_email"], [email], msg.as_string())
                    log.info("  ✓ Sent to %s", email)
                    sent += 1
                    state["sent_index"] += 1
                    save_state(cfg["state_file"], state)
                    if email != batch[-1]:
                        time.sleep(cfg["delay_emails"])
                except smtplib.SMTPException as e:
                    log.warning("  ✗ Failed %s: %s", email, e)

                if progress_callback:
                    progress_callback(idx + 1, total_in_batch)

    except Exception as e:
        log.error("SMTP connection error: %s", e)
        return 0

    return sent
